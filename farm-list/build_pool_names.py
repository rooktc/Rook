#!/usr/bin/env python3
"""One-time seed of pool_names.json: DefiLlama pool id -> display names.

The free API lacks the site's display names ("Pool" = receipt token like
aEthweETH, "Pool Meta" = e.g. 'Aave Ethereum weETH', and token-cased
references like 'weETH'). Seed them from the user's site CSV exports:
  1. exact ids from rows whose previous sheet build carried a DefiLlama link
  2. remaining rows matched to the live pool list on
     (protocol, chain, symbol) + nearest TVL (unambiguous matches only)
"""
import csv, json, sys
import openpyxl

SP = '/tmp/claude-0/-home-user-Rook/1e677254-eaa2-5dfc-9516-9447487826f9/scratchpad'
U = '/root/.claude/uploads/1e677254-eaa2-5dfc-9516-9447487826f9'
sys.path.insert(0, SP)
from refresh_farm_list import PROTOCOLS

pools = json.load(open(f'{SP}/pools.json'))['data']
mapping = {}

# --- source rows from the Aug-3 site CSVs (Pool, Reference, Pool Meta are display-cased)
csv_rows = []
for f in ['f67aa356-yieldpools.csv', 'ec7ab09d-yieldpools_2.csv']:
    csv_rows += list(csv.DictReader(open(f'{U}/{f}')))

# --- 1) exact ids via hyperlinks in the Aug-3 sheet build
def num(x):
    try: return float(x)
    except Exception: return None

linked = {}
wb = openpyxl.load_workbook(f'{SP}/farm_list_update_2026-08-03.xlsx')
for tab in ['USD Farms', 'ETH Farms']:
    ws = wb[tab]
    r = 4
    while ws.cell(r, 1).value:
        link = ws.cell(r, 6).hyperlink
        if link and '/pool/' in link.target:
            pid = link.target.rsplit('/', 1)[1]
            mapping[pid] = dict(pool=ws.cell(r, 17).value, farm=ws.cell(r, 5).value,
                                ref=ws.cell(r, 7).value)
            linked[(ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 17).value)] = pid
        r += 1
print('from links:', len(mapping))

# --- 2) TVL-nearest match for the rest
api_by_key = {}
for p in pools:
    api_by_key.setdefault((p['project'], p['chain'], p['symbol'].upper()), []).append(p)

ambig = unmatched = 0
for row in csv_rows:
    slug = PROTOCOLS.get(row['Protocol'])
    if not slug:
        continue
    key = (slug, row['Chain'], row['Reference'].upper())
    cands = api_by_key.get(key, [])
    tvl, apy = num(row['TVL']), num(row['APY'])
    if not cands or not tvl:
        unmatched += 1
        continue
    if len(cands) == 1:
        best = cands[0]
        ratio = (best['tvlUsd'] or 0) / tvl
        if not (0.2 <= ratio <= 5):
            unmatched += 1
            continue
    else:
        # multiple pools share (protocol, chain, symbol): combine relative TVL
        # distance and APY distance (2-day-old snapshot, still discriminative)
        def dist(p):
            dt = abs((p['tvlUsd'] or 0) - tvl) / max(tvl, 1)
            da = abs((p['apy'] or 0) - apy) / max(abs(apy), 1) if apy is not None else 0
            return dt + da
        scored = sorted(cands, key=dist)
        best = scored[0]
        if dist(best) > 0.8 or dist(scored[1]) < 1.5 * dist(best) + 0.05:
            ambig += 1
            continue
    if best['pool'] not in mapping:
        mapping[best['pool']] = dict(pool=row['Pool'], farm=row['Pool Meta'] or row['Pool'],
                                     ref=row['Reference'])

print(f'total mapped: {len(mapping)}  (csv rows {len(csv_rows)}, ambiguous {ambig}, unmatched {unmatched})')
json.dump(mapping, open(f'{SP}/pool_names.json', 'w'), indent=0)
