#!/usr/bin/env python3
"""Dump the USD/ETH/Looping tabs of a built workbook to feed CSVs for the
Apps Script updater installed in the original Google Sheet.

Layout: row 1 = title line (goes to A1), row 2 = headers, rows 3+ = data.
Farm tabs carry extra trailing LinkURL and Flags columns (Flags from the
workbook's hidden column S). The Looping feed carries computed Max ROE
values (the updater script restores the formula in G).

Usage: make_feed_csvs.py <workbook.xlsx> <usd.csv> <eth.csv> [loop.csv] [check.csv]
"""
import csv
import sys

import openpyxl

wb = openpyxl.load_workbook(sys.argv[1])
for tab, out in [('USD Farms', sys.argv[2]), ('ETH Farms', sys.argv[3])]:
    ws = wb[tab]
    with open(out, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow([ws.cell(1, 1).value] + [''] * 19)
        w.writerow([ws.cell(3, c).value for c in range(1, 19)] + ['LinkURL', 'Flags'])
        r = 4
        n = 0
        while ws.cell(r, 1).value:
            row = [ws.cell(r, c).value for c in range(1, 19)]
            link = ws.cell(r, 6).hyperlink
            row.append(link.target if link else '')
            row.append(ws.cell(r, 19).value or '')
            w.writerow(['' if v is None else v for v in row])
            r += 1
            n += 1
    print(f'{tab}: {n} rows -> {out}')

if len(sys.argv) > 4:
    ws = wb['Looping']
    with open(sys.argv[4], 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow([ws.cell(1, 1).value] + [''] * 13)
        w.writerow([ws.cell(3, c).value for c in range(1, 15)])
        r = 4
        n = 0
        while ws.cell(r, 1).value:
            row = [ws.cell(r, c).value for c in range(1, 15)]
            f_, h, i = row[5], row[7], row[8]
            try:
                row[6] = round((h + i) * (f_ - 1) + h, 5)  # formula -> value for the feed
            except TypeError:
                row[6] = ''
            w.writerow(['' if v is None else v for v in row])
            r += 1
            n += 1
    print(f'Looping: {n} rows -> {sys.argv[4]}')

if len(sys.argv) > 5 and 'Data Check' in wb.sheetnames:
    ws = wb['Data Check']
    with open(sys.argv[5], 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow([ws.cell(1, 1).value] + [''] * 9)
        w.writerow([ws.cell(2, c).value for c in range(1, 11)])
        r = 3
        n = 0
        while ws.cell(r, 1).value:
            w.writerow(['' if ws.cell(r, c).value is None else ws.cell(r, c).value
                        for c in range(1, 11)])
            r += 1
            n += 1
    print(f'Data Check: {n} rows -> {sys.argv[5]}')
