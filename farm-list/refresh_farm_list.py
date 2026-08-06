#!/usr/bin/env python3
"""Daily refresh of the USD/ETH farm-list spreadsheet from the DefiLlama API.

Reproduces the sheet at
https://docs.google.com/spreadsheets/d/1LiEDhK1PtCZH-au9BgNfwlN4_DrnrRSTO8rx4bO0SP4
originally built from defillama.com/yields CSV exports:
  - filters: protocol whitelist, TVL >= $500k, APY <= 100%,
    stablecoin pools (USD tab) / ETH-family pools (ETH tab)
  - sheet rule: keep pools with APY >= 6% (USD) / >= 3% (ETH)
  - "intrinsic" APY (underlying LST/yield-token rate) added the same way the
    site's apyComponents=reward,intrinsic toggle does
  - 7d / 30d columns are trailing averages computed from each pool's daily
    APY history (yields.llama.fi/chart/<pool>)
  - stability rating/score are computed here (coefficient of variation over
    the 30d series) — approximates, but does not equal, the site's metric
  - Holders is not available via the public API and is left blank
  - the Looping tab is carried over from the template unchanged

Usage: python3 refresh_farm_list.py <template.xlsx> <out.xlsx>
"""
import collections
import copy
import datetime
import json
import re
import statistics
import sys
import time
import urllib.request

import openpyxl

API = 'https://yields.llama.fi'
MIN_TVL = 200_000
MAX_APY = 100.0
CUTOFF = {'USD': 3.0, 'ETH': 2.0}

# display name -> DefiLlama project slug, from the user's defillama.com URLs
PROTOCOLS = {
    'Uniswap V2': 'uniswap-v2', 'Uniswap V3': 'uniswap-v3', 'Raydium AMM': 'raydium-amm',
    'Uniswap V4': 'uniswap-v4', 'Orca DEX': 'orca-dex', 'Curve DEX': 'curve-dex',
    'Beefy': 'beefy', 'Aerodrome Slipstream': 'aerodrome-slipstream', 'Morpho Blue': 'morpho-blue',
    'Aave V3': 'aave-v3', 'Convex Finance': 'convex-finance', 'Pendle': 'pendle',
    'Stake DAO': 'stake-dao', 'Balancer V3': 'balancer-v3', 'Aerodrome V1': 'aerodrome-v1',
    'Kamino Lend': 'kamino-lend', 'Euler V2': 'euler-v2', 'Yearn Finance': 'yearn-finance',
    'Kamino Liquidity': 'kamino-liquidity', 'Balancer V2': 'balancer-v2',
    'Velodrome V2': 'velodrome-v2', 'Project X': 'project-x', 'Project 0': 'project-0',
    'Dolomite': 'dolomite', 'Fusion by IPOR': 'fusion-by-ipor', 'PancakeSwap AMM': 'pancakeswap-amm',
    'Velodrome V3': 'velodrome-v3', 'Fluid DEX': 'fluid-dex', 'Camelot V3': 'camelot-v3',
    'Camelot V2': 'camelot-v2', 'Moonwell Lending': 'moonwell-lending',
    'Curve LlamaLend': 'curve-llamalend', 'Curvance': 'curvance', 'fx Protocol': 'fx-protocol',
    'Concrete': 'concrete', 'Fluid Lending': 'fluid-lending', 'Ember Protocol': 'ember-protocol',
    'Silo V2': 'silo-v2', 'Midas RWA': 'midas-rwa', 'Loopscale': 'loopscale',
    'JustLend V1': 'justlend-v1', 'Royco V2': 'royco-v2', 'Scallop Lend': 'scallop-lend',
    'Lista Lending': 'lista-lending', 'Fraxlend': 'fraxlend', 'Compound V3': 'compound-v3',
    'Mento V3': 'mento-v3', 'Compound V2': 'compound-v2', 'HyperLend Pooled': 'hyperlend-pooled',
    'HypurrFi Pooled': 'hypurrfi-pooled', 'Neverland': 'neverland', 'SparkLend': 'sparklend',
    'TermMax': 'termmax', 'Current': 'current', 'Strata Markets': 'strata-markets',
    'Gearbox': 'gearbox', 'Nest Credit': 'nest-credit', 'Tydro': 'tydro', 'YieldNest': 'yieldnest',
    'D2 Finance': 'd2-finance', 'Gains Network': 'gains-network', 'Superform': 'superform',
    'Reserve Protocol': 'reserve-protocol', 'Sky Lending': 'sky-lending',
    'Spark Savings': 'spark-savings', 'Upshift': 'upshift', 'Yuzu Money': 'yuzu-money',
    'Makina': 'makina', 'Kai Finance': 'kai-finance', 'Pareto Credit': 'pareto-credit',
    'Yuzu Finance': 'yuzu-finance', 'ether.fi Stake': 'ether.fi-stake',
    'Multipli.fi': 'multipli-fi', 'Clearpool Lending': 'clearpool-lending',
    'ether.fi Liquid': 'ether.fi-liquid', 'Gauntlet': 'gauntlet', 'Veda': 'veda',
    'Liquity V2': 'liquity-v2', 'Maple': 'maple', 'Mystic Finance Lending': 'mystic-finance-lending',
    'Steakhouse Financial': 'steakhouse-financial', '3Jane Lending': '3jane-lending',
    'Falcon Finance': 'falcon-finance', 'Re': 're', 'Aave V4': 'aave-v4',
    'Fluid Lite': 'fluid-lite', 'Lagoon': 'lagoon', 'Joe V2.1': 'joe-v2.1', 'Joe V2.2': 'joe-v2.2',
}
SLUG2NAME = {v: k for k, v in PROTOCOLS.items()}

# yield-bearing tokens whose underlying rate the site adds as "intrinsic":
# symbol -> (issuer project slug, issuer pool symbol)
ISSUERS = {
    'WEETH': ('ether.fi-stake', 'WEETH'), 'WSTETH': ('lido', 'STETH'), 'STETH': ('lido', 'STETH'),
    'CBETH': ('coinbase-wrapped-staked-eth', 'CBETH'), 'RETH': ('rocket-pool', 'RETH'),
    'OSETH': ('stakewise-v2', 'OSETH'), 'EZETH': ('renzo', 'EZETH'), 'RSETH': ('kelp', 'RSETH'),
    'METH': ('meth-protocol', 'METH'), 'SUSDE': ('ethena-usde', 'SUSDE'),
    'SDAI': ('sdai', 'SDAI'), 'SUSDS': ('sky-lending', 'SUSDS'), 'SYRUPUSDC': ('maple', 'SYRUPUSDC'),
    'SFRXETH': ('frax-ether', 'SFRXETH'),
}

RATING_COLOR = {'stable': 'FF1E7145', 'mixed': 'FFB45F06', 'volatile': 'FFC00000'}
RISK_COLOR = {'Low': 'FF1E7145', 'Med-Low': 'FF7F8C1E', 'Med-High': 'FFB45F06',
              'High': 'FFC00000', 'Medium': 'FFB45F06', 'Med.': 'FFB45F06'}
RISK_RANK = {'Low': 0, 'Medium': 1, 'High': 2}
RISK_ALIAS = {'ETH+': 'ETHPLUS', 'USDT0': 'USDT', 'USD₮0': 'USDT', 'USD₮': 'USDT',
              'SCRVUSD': 'CRVUSD', 'SFRXUSD': 'FRXUSD', 'WFRXETH': 'FRXETH'}


def risk_norm(leg):
    leg = RISK_ALIAS.get(leg, RISK_ALIAS.get(leg.upper(), leg))
    return re.sub(r'[^A-Z0-9]', '', (leg or '').upper())


# composite risk score: weighted average of normalized source scores,
# 1-10 higher = safer; buckets 1-3 High, 4-7 Medium, 8-10 Low
SOURCE_WEIGHTS = {'tid': 0.30, 'pharos': 0.25, 'credora': 0.25, 'yearn': 0.20}
PHAROS_LETTER = {'A+': 9.5, 'A': 9.0, 'A-': 8.5, 'B+': 7.5, 'B': 7.0, 'B-': 6.5,
                 'C+': 5.5, 'C': 5.0, 'C-': 4.5, 'D': 3.0, 'F': 1.5}
# Credora's PD curve: A band = minimal expected loss, B moderate, C/D speculative
CREDORA_LETTER = {'A+': 9.5, 'A': 9.0, 'A-': 8.4, 'B+': 7.4, 'B': 6.6, 'B-': 5.8,
                  'C+': 4.8, 'C': 4.0, 'C-': 3.2, 'D+': 2.2, 'D': 1.5}
YEARN_ANCHORS = [(1.0, 10.0), (2.5, 8.0), (3.5, 4.0), (5.0, 1.0)]


def entry_score10(e):
    """Normalize a source entry to the 1-10 higher-is-safer scale."""
    if e['source'] == 'tid':
        return float(e['score'])
    if e['source'] == 'pharos':
        if e.get('num') is not None:
            return max(0.5, min(10.0, e['num'] / 10))
        return PHAROS_LETTER.get(str(e['score']).upper())
    if e['source'] == 'credora':
        return CREDORA_LETTER.get(str(e['score']).upper())
    # yearn 1-5 lower = safer: piecewise-linear so their labels align with
    # the 8/4 bucket edges
    s = float(e['score'])
    for (x1, y1), (x2, y2) in zip(YEARN_ANCHORS, YEARN_ANCHORS[1:]):
        if s <= x2:
            s = max(x1, s)
            return y1 + (s - x1) * (y2 - y1) / (x2 - x1)
    return 1.0


def asset_score(leg, risk_scores):
    """Weighted composite for one normalized symbol -> (score, parts) or None."""
    per_source = {}
    for e in risk_scores.get(leg, []):
        v = entry_score10(e)
        if v is not None:
            per_source.setdefault(e['source'], []).append(v)
    if not per_source:
        return None
    wsum = sum(SOURCE_WEIGHTS[s] for s in per_source)
    score = sum(SOURCE_WEIGHTS[s] * (sum(vs) / len(vs))
                for s, vs in per_source.items()) / wsum
    parts = ', '.join(f'{s} {sum(vs)/len(vs):.1f}' for s, vs in sorted(per_source.items()))
    return score, parts


# collateral-quality baseline for major non-stable collateral the
# stablecoin/vault-focused risk sources never rate: liquidity depth, oracle
# robustness and liquidation behavior, 1-10 higher = safer. Used ONLY for
# the exposure leg — never as a base-asset score.
COLLATERAL_BASELINE = {
    'WETH': 9.0, 'ETH': 9.0, 'WSTETH': 8.5, 'STETH': 8.5, 'CBETH': 8.3,
    'RETH': 8.3, 'WEETH': 8.0, 'EZETH': 7.5, 'RSETH': 7.5, 'METH': 7.8,
    'SFRXETH': 8.0, 'OSETH': 7.8,
    'WBTC': 8.5, 'CBBTC': 8.5, 'TBTC': 7.8, 'LBTC': 7.5, 'FBTC': 7.3,
    'BTC': 8.5, 'XAUT': 7.5, 'PAXG': 7.5,
    'SOL': 8.5, 'JITOSOL': 8.0, 'MSOL': 8.0, 'JUPSOL': 7.8, 'BSOL': 7.8,
    'BNSOL': 7.5, 'INF': 7.3, 'DSOL': 7.3, 'HSOL': 7.3, 'VSOL': 7.0,
    'BNB': 8.5, 'WBNB': 8.5, 'SLISBNB': 7.5, 'SLISBNBX': 7.0,
    'SUI': 8.0, 'WSUI': 8.0, 'HASUI': 7.5, 'AFSUI': 7.5, 'VSUI': 7.3,
    'AVAX': 8.0, 'WAVAX': 8.0, 'POL': 7.5, 'WMATIC': 7.5,
    'ARB': 7.5, 'OP': 7.5, 'LINK': 8.0, 'UNI': 7.5, 'AAVE': 8.0,
    'MKR': 7.5, 'SKY': 7.0, 'LDO': 7.0, 'CRV': 6.5, 'PENDLE': 6.5,
    'HYPE': 7.0, 'WHYPE': 7.0, 'MON': 6.5, 'WMON': 6.5, 'SHMON': 6.0,
    'JLP': 6.5, 'JTO': 6.8, 'JUP': 6.8, 'ENA': 6.5, 'EIGEN': 6.5,
}


def _collat_score(sym, risk_scores):
    """Score one collateral symbol: rated sources, then the baseline table,
    then a PT token's underlying (with a wrapper haircut)."""
    n = risk_norm(sym)
    s = asset_score(n, risk_scores)
    if s is not None:
        return s[0], f'{sym} {s[0]:.1f}'
    if n in COLLATERAL_BASELINE:
        v = COLLATERAL_BASELINE[n]
        return v, f'{sym} {v:.1f}*'
    if sym.upper().startswith('PT-') and sym.count('-') >= 1:
        u = sym.split('-')[1]
        un = risk_norm(u)
        s = asset_score(un, risk_scores)
        v = s[0] if s is not None else COLLATERAL_BASELINE.get(un)
        if v is not None:
            v = max(0.5, v - 0.5)   # maturity/oracle wrapper haircut
            return v, f'{sym}→{u} {v:.1f}'
    return None, None


def exposure_leg(coll, risk_scores):
    """Weighted safety score over an exposure list [(symbol, weight), ...].

    Sizable exposure (>=10%) to an asset nothing rates is itself the risk
    signal — it scores 3.0 rather than being skipped. Returns None when
    less than half of the book is visible.
    """
    tw = sum(item[1] for item in coll)
    if tw <= 0:
        return None
    num, den, parts = 0.0, 0.0, []
    for item in sorted(coll, key=lambda x: -x[1])[:10]:
        sym, w = item[0], item[1]
        share = w / tw
        if len(item) > 2 and item[2] is not None:   # pre-scored venue entry
            v, label = item[2], f'{sym} {item[2]:.1f}†'
        else:
            v, label = _collat_score(sym, risk_scores)
        if v is not None:
            num += v * w
            den += w
            parts.append(f'{label}@{share:.0%}')
        elif share >= 0.10:
            num += 3.0 * w
            den += w
            parts.append(f'{sym} unrated@{share:.0%}')
    if den < 0.5 * tw:
        return None
    return num / den, ', '.join(parts)


# chain maturity: points deducted for settlement/infra/oracle immaturity.
# Unlisted chains take the 0.5 default.
CHAIN_TIER = {'Ethereum': 0.0,
              'Base': 0.25, 'Arbitrum': 0.25, 'OP Mainnet': 0.25,
              'Polygon': 0.25, 'BSC': 0.25, 'Avalanche': 0.25,
              'Gnosis': 0.25, 'Solana': 0.25, 'Linea': 0.25, 'Scroll': 0.25,
              'Sui': 0.5, 'Tron': 0.5, 'Mantle': 0.5, 'Fraxtal': 0.5,
              'Sonic': 0.75, 'Monad': 0.75, 'Plasma': 0.75, 'Katana': 0.75,
              'Hyperliquid L1': 0.75, 'HyperEVM': 0.75, 'Unichain': 0.75,
              'Etherlink': 0.75, 'TAC': 0.75, 'XRPL EVM': 0.75, 'RSK': 0.75,
              'World Chain': 0.75}
TRUSTED_ORACLES = {'chainlink', 'chronicle', 'redstone', 'pyth', 'api3',
                   'lido', 'oval', 'compound', 'uniswap', 'pendle'}
_LEVERED = re.compile(r'lo{2,}p|\blever', re.IGNORECASE)

# Vault taxonomy. Strategy vaults are curated allocators — risk lives in
# the manager's book; lending vaults are passive deposit pools — risk lives
# in borrower collateral/utilization/LTV. Either kind caps at 5.9 when the
# relevant exposure is not visible from any source.
STRATEGY_VAULTS = {'Fusion by IPOR', 'Lagoon', 'Ember Protocol', 'Upshift',
                   'Makina', 'Concrete', 'Superform', 'Veda', 'Gauntlet',
                   'ether.fi Liquid', 'D2 Finance', 'Multipli.fi', 'Re',
                   'Nest Credit', 'Tydro', 'Kai Finance', 'Pareto Credit',
                   '3Jane Lending', 'Royco V2', 'Steakhouse Financial',
                   'Yearn Finance', 'Yuzu Money', 'Yuzu Finance'}
LENDING_OPAQUE = {'Loopscale', 'Current', 'Project 0', 'TermMax', 'Gearbox',
                  'Clearpool Lending', 'Scallop Lend'}


_PT_DATE = re.compile(r'(\d{1,2})([A-Z]{3})(\d{4})')
_MONTHS = {'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
           'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12}


def pt_tenor_days(row):
    """Days to maturity parsed from the PT name (e.g. PT-sUSDe-13AUG2026)."""
    m = _PT_DATE.search(f"{row.get('farm_disp') or ''} {row.get('meta') or ''}".upper())
    if not m or m.group(2) not in _MONTHS:
        return None
    try:
        mat = datetime.date(int(m.group(3)), _MONTHS[m.group(2)], int(m.group(1)))
    except ValueError:
        return None
    return max((mat - datetime.date.today()).days, 0)


def compute_risk(row, risk_scores):
    """Composite High/Medium/Low label from a weighted 1-10 safety score.

    Legs: per base asset (LP rows take the worst leg) plus an exposure leg —
    the debt-weighted collateral of a lending market, or a vault's actual
    allocation — captured during verification. The overall score is the
    worst leg. Row signals deduct points; utilization >92% deducts (deposits
    may not be withdrawable); a row that is not VERIFIED cannot reach Low,
    and an opaque curated vault cannot score above 6.9.
    Buckets: >=8 Low, >=4 Medium, else High.
    """
    # risk_legs overrides the symbol split when the deposit ticker is not the
    # real exposure (Midas mTokens, Strata tranche underlyings)
    leg_scores, basis, unrated_legs = {}, [], []
    for leg_raw in (row.get('risk_legs') or row['symbol'].split('-')):
        if not leg_raw:
            continue
        leg = risk_norm(leg_raw)
        s = asset_score(leg, risk_scores)
        if s is not None:
            leg_scores[leg] = s[0]
            basis.append(f'{leg} {s[0]:.1f} ({s[1]})')
            continue
        # baseline/PT fallback so majors (WETH, wstETH, SOL...) still score
        v, label = _collat_score(leg_raw, risk_scores)
        if v is not None:
            leg_scores[leg] = v
            basis.append(label)
        else:
            unrated_legs.append(leg)
    # an LP whose other leg nothing rates: the unknown leg IS the risk —
    # it enters at 3.0 instead of being silently ignored
    if leg_scores and unrated_legs:
        for leg in unrated_legs:
            leg_scores[leg] = 3.0
            basis.append(f'{leg} unrated 3.0')

    exp = None
    if row.get('collat'):
        exp = exposure_leg(row['collat'], risk_scores)
        if exp:
            kind = row.get('collat_kind') or 'collateral'
            basis.append(f'{kind} leg {exp[0]:.1f} ({exp[1]})')
    cred = row.get('credora_vault')
    cred_score = CREDORA_LETTER.get(str((cred or {}).get('rating')).upper()) if cred else None
    if cred_score is not None:
        leg_scores['CREDORA-VAULT'] = cred_score
        basis.append(f"credora vault {cred['rating']} "
                     f"(psl {100 * (cred.get('psl') or 0):.2f}%)")

    flags = row.get('flags') or []
    red = any(f.startswith(RED_FLAGS) for f in flags)
    if leg_scores or exp:
        cands = {}
        if leg_scores:
            worst_leg = min(leg_scores, key=leg_scores.get)
            cands['asset ' + worst_leg] = leg_scores[worst_leg]
            if len(leg_scores) > 1:
                basis.append(f'worst leg: {worst_leg}')
        if exp:
            cands[(row.get('collat_kind') or 'collateral') + ' leg'] = exp[0]
        worst = min(cands, key=cands.get)
        score = cands[worst]
        if len(cands) > 1 and not worst.startswith('asset'):
            basis.append(f'{worst} governs')
        deductions = []
        if red:
            deductions.append(('red flags', 2.0))
        if 'SELF-REPORTED(flat 30d history)' in flags:
            deductions.append(('self-reported', 2.0))
        if (row.get('tvl') or 0) < 1_000_000:
            deductions.append(('tvl<1M', 1.0))
        if (row['name'] == 'Strata Markets'
                and row['symbol'].upper().startswith('JR')):
            deductions.append(('junior tranche (first loss)', 1.5))
        if row['name'] == 'Pendle' and 'buying PT' in (row.get('meta') or ''):
            deductions.append(('PT wrapper', 0.5))
        ftype = row.get('ftype') or ''
        if ftype not in ('LP', 'PT', 'LP (Pendle)'):
            # lending/vault signals: liquidation margin, oracle, utilization
            lltv = row.get('lltv')
            if lltv is not None:
                if lltv >= 0.965:
                    deductions.append((f'lltv {lltv:.0%}', 1.0))
                elif lltv >= 0.945:
                    deductions.append((f'lltv {lltv:.0%}', 0.5))
            orc = row.get('oracle')
            if orc is not None and not (set(p.lower() for p in orc) & TRUSTED_ORACLES):
                deductions.append(('exotic oracle', 0.5))
            util = row.get('util')
            if util is not None:
                u = util if util <= 1.5 else util / 100
                if u > 0.92:
                    deductions.append((f'util {u:.0%}', 1.0))
        if _LEVERED.search(f"{row.get('farm_disp') or ''} {row.get('meta') or ''}"):
            deductions.append(('leveraged strategy', 1.0))
        if ftype == 'PT':
            # fixed-rate position: longer tenor = more rate/exit risk, and
            # exiting early depends on the AMM's liquidity
            days = pt_tenor_days(row)
            if days is not None:
                if days > 365:
                    deductions.append((f'tenor {days}d', 1.5))
                elif days > 180:
                    deductions.append((f'tenor {days}d', 1.0))
                elif days > 90:
                    deductions.append((f'tenor {days}d', 0.5))
            if (row.get('tvl') or 0) < 5_000_000:
                deductions.append(('PT liquidity <5M', 0.5))
        if (row.get('collat_kind') in ('allocation', 'strategy book')
                and row.get('collat')):
            tw = sum(i[1] for i in row['collat'])
            top = max(i[1] for i in row['collat'])
            if tw > 0 and top / tw >= 0.90:
                deductions.append((f'concentrated book {top / tw:.0%}', 0.5))
        imb = row.get('pool_imb')
        if imb is not None:
            # skew vs equal weight: LPs in a lopsided pool effectively hold
            # the heavy (usually weaker) asset
            ratio, sym, share = imb
            # share tier catches extreme 2-coin skews (their ratio caps at 2x)
            if share >= 0.85 or ratio >= 2.2:
                deductions.append((f'pool {share:.0%} in {sym}', 1.0))
            elif ratio >= 1.5:
                deductions.append((f'pool {share:.0%} in {sym}', 0.5))
        for name, pts in deductions:
            score -= pts
            basis.append(f'-{pts:g} {name}')
        if (row['name'] == 'Strata Markets'
                and row['symbol'].upper().startswith('SR')):
            score += 0.5   # loss-absorbing junior tranche beneath it
            basis.append('+0.5 senior tranche buffer')
        if score >= 8 and row.get('verdict') != 'VERIFIED':
            score = 7.9
            basis.append('capped 7.9: not VERIFIED')
        # a vault token the risk sources rate directly (yBOLD, syrupUSDC)
        # carries a product-level assessment — but a plain deposit currency
        # (USDC, WETH...) rating says nothing about the vault
        PLAIN = {'USDC', 'USDT', 'USDT0', 'DAI', 'USDS', 'USDE', 'WETH', 'ETH',
                 'WBTC', 'CBBTC', 'GHO', 'FRAX', 'FRXUSD', 'LUSD', 'CRVUSD',
                 'BOLD', 'RLUSD', 'PYUSD', 'USDP', 'TUSD', 'USD1', 'AUSD',
                 'USDG', 'EURC', 'SUI', 'SOL', 'BNB', 'MON', 'HYPE', 'XAUT',
                 'WSTETH', 'STETH', 'WEETH', 'CBETH', 'RETH', 'USDAI', 'USN'}
        n_sym = risk_norm(row['symbol'])
        product_rated = (cred_score is not None
                         or (n_sym not in PLAIN and '-' not in row['symbol']
                             and asset_score(n_sym, risk_scores) is not None))
        if score > 5.9 and not row.get('collat') and not product_rated:
            # the deposit asset's score says nothing about exposure we
            # cannot see — Med-High is the ceiling either way. Exception:
            # a vault token the risk sources rate directly (yBOLD, syrupUSDC)
            # already carries a product-level assessment.
            if row['name'] in STRATEGY_VAULTS:
                score = 5.9
                basis.append('capped 5.9: allocation not visible')
            elif row['name'] in LENDING_OPAQUE:
                score = 5.9
                basis.append('capped 5.9: collateral not visible')
        score = max(0.5, min(10.0, score))
        risk = ('Low' if score >= 8 else 'Med-Low' if score >= 6 else
                'Med-High' if score >= 4 else 'High')
        basis.insert(0, f'score {score:.1f}')
        return risk, '; '.join(basis), round(score, 1)
    # unrated assets: label from row signals only. An asset nothing rates
    # whose yield also failed verification defaults to High — unknown and
    # unconfirmed is not a middle case.
    weak = (red or 'SELF-REPORTED(flat 30d history)' in flags
            or row.get('rating') == 'volatile' or (row.get('tvl') or 0) < 1_000_000
            or row.get('verdict') not in ('VERIFIED', 'CAUTION'))
    risk = 'High' if weak else 'Med-High'
    basis.append('unrated: internal signals only')
    return risk, '; '.join(basis), None


UA = 'Mozilla/5.0 (X11; Linux x86_64) farm-list-verifier'


def get(url, retries=4):
    for i in range(retries):
        try:
            req = urllib.request.Request(url, headers={'User-Agent': UA})
            with urllib.request.urlopen(req, timeout=60) as r:
                return json.load(r)
        except Exception:
            if i == retries - 1:
                raise
            time.sleep(2 ** i)


import re

_ETH_PART = re.compile(r'ETH([+XYSB2]|\.[EB])?$')  # ETH, wstETH, ETHx, ETH+, weETHs, WETH.e ...


def is_eth_family(symbol):
    # every hyphen-separated leg must be an ETH-denominated token; rejects
    # tickers that merely contain "ETH" (ETHFI, ETHENA, ...)
    parts = [p for p in symbol.split('-') if p]
    return bool(parts) and all(_ETH_PART.search(p.upper()) for p in parts)


def farm_type(name, symbol, farm, meta):
    if name == 'Pendle':
        # API poolMeta: 'For buying PT-...' vs 'For LP | Maturity ...'
        is_pt = (farm or '').startswith('PT ') or 'buying PT' in (meta or '')
        return 'PT' if is_pt else 'LP (Pendle)'
    if name in ('Beefy', 'Convex Finance'):
        return 'LP'
    return 'LP' if '-' in symbol else 'Lending'


def main(template, outfile, names_file=None, loops_file=None):
    pools = get(f'{API}/pools')['data']
    print(f'{len(pools)} pools fetched')
    names = {}
    if names_file:
        try:
            names = json.load(open(names_file))
        except Exception as e:
            print('pool names file not loaded:', e)

    # intrinsic rates: issuer pool's current APY (largest instance)
    intr_rate = {}
    for sym, (proj, psym) in ISSUERS.items():
        cands = [p for p in pools if p['project'] == proj and p['symbol'].upper() == psym
                 and p.get('exposure') == 'single']
        if cands:
            intr_rate[sym] = max(cands, key=lambda p: p['tvlUsd'])['apy'] or 0
    print('intrinsic rates:', {k: round(v, 3) for k, v in intr_rate.items()})

    tabs = {'USD Farms': [], 'ETH Farms': []}
    charts_needed = []
    for p in pools:
        name = SLUG2NAME.get(p['project'])
        if not name or (p['tvlUsd'] or 0) < MIN_TVL:
            continue
        sym = p['symbol']
        if p.get('stablecoin'):
            tab = 'USD Farms'
        elif is_eth_family(sym):
            tab = 'ETH Farms'
        else:
            continue
        intr = None
        if sym.upper() in intr_rate and p.get('exposure') == 'single' \
                and p['project'] != ISSUERS[sym.upper()][0]:
            intr = intr_rate[sym.upper()]
        base, rew = p.get('apyBase'), p.get('apyReward')
        now = None
        if base is not None or rew is not None or intr is not None:
            now = (base or 0) + (rew or 0) + (intr or 0)
        elif p.get('apy') is not None:
            now = p['apy']
        mean30 = (p.get('apyMean30d') or 0) + (intr or 0)
        cut = CUTOFF[tab.split()[0]]
        # candidates: current APY qualifies, or no current APY but the 30d
        # trailing average qualifies (matches the old sheet's coalesce rule)
        if now is not None and cut <= now <= MAX_APY:
            pass
        elif now in (None, 0) and cut <= mean30 <= MAX_APY:
            pass
        else:
            continue
        row = dict(tab=tab, name=name, chain=p['chain'], symbol=sym, meta=p.get('poolMeta'),
                   tvl=p['tvlUsd'], now=now, base=base, rew=rew, intr=intr,
                   pool_id=p['pool'], d7=None, d30=round(mean30, 5) if mean30 else None,
                   rating=None, score=None,
                   _underlying=(p.get('underlyingTokens') or [None])[0],
                   _underlyings=p.get('underlyingTokens') or [])
        tabs[tab].append(row)
        charts_needed.append(row)

    print(f"candidates: USD {len(tabs['USD Farms'])}, ETH {len(tabs['ETH Farms'])};"
          f" fetching {len(charts_needed)} charts")
    def enrich(row):
        try:
            hist = get(f"{API}/chart/{row['pool_id']}")['data']
        except Exception as e:
            print('  chart failed', row['pool_id'], e)
            return
        raw = [d['apy'] for d in hist[-30:] if d.get('apy') is not None]
        row['_raw'] = raw
        row['_last_ts'] = hist[-1]['timestamp'] if hist else None
        series = [x + (row['intr'] or 0) for x in raw]
        if len(series) >= 7:
            row['d7'] = round(statistics.mean(series[-7:]), 5)
            row['d30'] = round(statistics.mean(series), 5)
        if len(series) >= 10:
            m = statistics.mean(series)
            cv = statistics.pstdev(series) / m if m > 0 else 1.0
            score = 1 / (1 + 3 * cv)
            row['score'] = round(score, 5)
            row['rating'] = 'stable' if score >= 2 / 3 else ('mixed' if score >= 1 / 3 else 'volatile')

    import concurrent.futures
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
        done = 0
        for _ in ex.map(enrich, charts_needed):
            done += 1
            if done % 100 == 0:
                print(f'  {done}/{len(charts_needed)} charts')

    # final filter + ordering (current APY desc, then blank-now rows by 7d/30d)
    for tab, cut in [('USD Farms', CUTOFF['USD']), ('ETH Farms', CUTOFF['ETH'])]:
        rows = [r for r in tabs[tab]
                if (r['now'] is not None and cut <= r['now'] <= MAX_APY)
                or (not r['now'] and cut <= (r['d7'] or r['d30'] or 0) <= MAX_APY)]
        rows.sort(key=lambda r: (0 if r['now'] else 1,
                                 -(r['now'] if r['now'] else (r['d7'] or r['d30'] or 0))))
        tabs[tab] = rows

    registry, risk_scores = {}, {}
    if names_file:
        import os
        base = os.path.dirname(os.path.abspath(names_file))
        try:
            registry = json.load(open(os.path.join(base, 'vault_registry.json')))
            print(f'vault registry: {len(registry)} entries')
        except Exception:
            pass
        try:
            risk_scores = json.load(open(os.path.join(base, 'risk_scores.json')))
            print(f'risk scores: {len(risk_scores)} assets')
        except Exception:
            pass
    _MV_BOOKS['risk'] = risk_scores
    # display names before verification: some cross-checks join on them
    for rows in tabs.values():
        for r in rows:
            nm = names.get(r['pool_id'], {})
            r['farm_disp'] = nm.get('farm') or r['meta'] or nm.get('pool') or r['symbol']
    try:
        run_verification(tabs, registry)
    except Exception as e:
        print('verification skipped entirely:', e)
        for rows in tabs.values():
            for r in rows:
                r.setdefault('flags', [])
                r.setdefault('verdict', 'UNVERIFIED')
                r.setdefault('src_val', None)
                r.setdefault('src_name', None)

    risk_counts = collections.Counter()
    for rows in tabs.values():
        for r in rows:
            # score the real exposure, not the deposit ticker: Midas rows
            # carry the deposit asset as symbol with the mToken in meta;
            # Strata tranches wrap their underlying
            r['ftype'] = farm_type(r['name'], r['symbol'],
                                   r.get('farm_disp'), r.get('meta'))
            cv = risk_scores.get('_credora_vaults') or {}
            if cv and not r.get('credora_vault'):
                cid = next((k for k, v in CHAIN_IDS.items() if v == r['chain']), None)
                addrs = list(r.get('_pool_addrs') or [])
                ent = registry.get(r['pool_id'])
                if ent and ent.get('address'):
                    addrs.append(ent['address'])
                for a in addrs:
                    hit = cv.get(f'{cid}:{str(a).lower()}')
                    if hit:
                        r['credora_vault'] = hit
                        break
            if r['ftype'] in ('LP', 'LP (Pendle)'):
                # LP risk = the pooled assets themselves; lending-style
                # exposure legs don't apply
                r.pop('collat', None)
                r.pop('collat_kind', None)
            if r['name'] == 'Midas RWA' and r.get('meta'):
                r['risk_legs'] = [r['meta']]
            elif (r['name'] == 'Strata Markets'
                  and r['symbol'].upper()[:2] in ('SR', 'JR')):
                r['risk_legs'] = [r['symbol'][2:]]
            try:
                r['risk'], r['risk_basis'], r['risk_score'] = compute_risk(r, risk_scores)
            except Exception:
                r['risk'], r['risk_basis'], r['risk_score'] = None, '', None
            risk_counts[r['risk']] += 1
    print('risk labels:', dict(risk_counts))

    # ---- write workbook, preserving the template's styling ----
    wb = openpyxl.load_workbook(template)
    today = datetime.date.today().strftime('%-d %b %Y')
    for tab, rows in tabs.items():
        ws = wb[tab]
        old_last = max(r for r in range(4, ws.max_row + 1)
                       if ws.cell(r, 1).value not in (None, '')) if ws.max_row >= 4 else 3
        tmpl = {c: copy.copy(ws.cell(4, c)._style) for c in range(1, 19)}
        base_font = copy.copy(ws.cell(4, 1).font)
        restructure_farm_header(ws)
        for r in range(4, old_last + 1):
            for c in range(1, 21):
                cell = ws.cell(r, c)
                cell.value = None
                cell.hyperlink = None
                cell.style = 'Normal'
            if r in ws.row_dimensions:
                del ws.row_dimensions[r]
        asset = tab.split()[0]
        unnamed = 0
        for i, d in enumerate(rows):
            r = 4 + i
            ws.row_dimensions[r].height = 15.0
            nm = names.get(d['pool_id'], {})
            if not nm:
                unnamed += 1
            ref = nm.get('ref') or d['symbol']
            farm = nm.get('farm') or d['meta'] or nm.get('pool') or d['symbol']
            pool_name = nm.get('pool') or d['symbol']
            vals = [asset, d['name'], d['chain'],
                    farm_type(d['name'], ref, farm, d['meta']),
                    farm, d.get('risk'), 'DefiLlama', ref,
                    round(d['tvl']), round(d['now'], 5) if d['now'] is not None else None,
                    d['d7'], d['d30'],
                    round(d['base'], 5) if d['base'] is not None else None,
                    round(d['rew'], 5) if d['rew'] is not None else None,
                    round(d['intr'], 5) if d['intr'] is not None else None,
                    d.get('risk_basis') or '']
            # style source per new column: A..E keep 1..5, F(risk) borrows A,
            # G..O take the template's old F..N styles, Risk Note borrows
            # the old pool-name text style
            style_src = [1, 2, 3, 4, 5, 1, 6, 7, 8, 9, 10, 11, 12, 13, 14, 17]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(r, c)
                cell._style = copy.copy(tmpl[style_src[c - 1]])
                cell.value = v
            risk_cell = ws.cell(r, 6)
            risk_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='bottom')
            if d.get('risk') in RISK_COLOR:
                risk_cell.font = openpyxl.styles.Font(
                    name='Arial', size=10, bold=True, color=RISK_COLOR[d['risk']])
            link = ws.cell(r, 7)
            link.hyperlink = f"https://defillama.com/yields/pool/{d['pool_id']}"
            f = copy.copy(link.font)
            f.color = openpyxl.styles.Color(rgb='FF0563C1')
            f.underline = 'single'
            link.font = f
            link.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='bottom')
            note = ws.cell(r, 16)
            note.font = openpyxl.styles.Font(name='Arial', size=8, color='FF666666')
            note.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='bottom')
            d['farm_disp'] = farm
            # verification flags: highlight the APY (Now) cell, detail in hidden col Q
            flags = d.get('flags') or []
            if flags:
                shade = FLAG_FILL['red' if any(x.startswith(RED_FLAGS) for x in flags) else 'orange']
                ws.cell(r, 10).fill = openpyxl.styles.PatternFill('solid', fgColor=shade)
                ws.cell(r, 17).value = '; '.join(flags)
        ws.column_dimensions['Q'].hidden = True
        last = 3 + len(rows)
        ws.cell(1, 1).value = (f"{asset} Farm List  ·  DefiLlama API snapshot {today}"
                               f"  ·  {len(rows)} pools")
        ws.auto_filter.ref = f'A3:P{last}'
        ws.freeze_panes = 'C4'
        print(f'{tab}: {len(rows)} rows written ({unnamed} without carried display names)')

    try:
        build_check_tab(wb, tabs, today)
    except Exception as e:
        print('Data Check tab skipped:', e)

    if loops_file:
        try:
            loops = json.load(open(loops_file))
        except Exception as e:
            print('looping feed not loaded, tab carried over unchanged:', e)
            loops = None
        if loops is not None:
            if len(loops) < 50:
                print(f'looping feed too small ({len(loops)} rows), tab carried over unchanged')
            else:
                build_looping(wb, loops, today, risk_scores)

    wb.save(outfile)
    print('saved', outfile)


def loop_risk(d, risk_scores):
    """Risk for a leveraged loop: worst of deposit asset, borrow asset and
    venue quality, minus leverage/LLTV/utilization/exit-liquidity points.
    A depeg on either side of the loop liquidates the position."""
    basis, legs = [], []
    for sym in (d['depositAsset'], d['borrowAsset']):
        v, label = _collat_score(sym, risk_scores)
        if v is None:
            v, label = 3.0, f'{sym} unrated 3.0'
        legs.append(v)
        basis.append(label)
    venue = next((sc for k, sc in BOOK_VENUES if k in (d.get('protocol') or '').lower()), None)
    if venue is not None:
        legs.append(venue)
        basis.append(f"{d['protocol']} {venue:.1f}†")
    score = min(legs)
    lev = d.get('lev') or 1
    for cut, pts in ((10, 1.5), (5, 1.0), (3, 0.5)):
        if lev >= cut:
            score -= pts
            basis.append(f'-{pts:g} leverage {lev:g}x')
            break
    lltv = (d.get('lltv') or 0) / 100
    if lltv >= 0.965:
        score -= 1.0
        basis.append(f'-1 lltv {lltv:.0%}')
    elif lltv >= 0.945:
        score -= 0.5
        basis.append(f'-0.5 lltv {lltv:.0%}')
    if (d.get('utilization') or 0) > 92:
        score -= 1.0
        basis.append(f"-1 util {d['utilization']:.0f}%")
    if (d.get('liquidity') or 0) < 1_000_000:
        score -= 1.0
        basis.append('-1 exit liquidity <1M')
    score = max(0.5, min(10.0, score))
    risk = ('Low' if score >= 8 else 'Med-Low' if score >= 6 else
            'Med-High' if score >= 4 else 'High')
    return risk, f'score {score:.1f}; ' + '; '.join(basis) + f"; yieldz label: {d.get('risk')}"


def build_looping(wb, loops, today, risk_scores=None):
    """Rewrite the Looping tab from scraped yieldz.io/leverage rows.

    Keeps the old sheet's semantics: positive-carry loops only
    (deposit APY + borrow APY > 0), ETH block first then USD, each sorted
    by max leverage desc. Max leverage is derived from the displayed max
    ROE via the sheet's own formula: ROE = (H+I)*(F-1)+H  =>  F.
    """
    seen, rows = set(), []
    for r in loops:
        k = json.dumps(r, sort_keys=True)
        if k in seen:
            continue
        seen.add(k)
        dep, bor = r.get('depApy'), r.get('borApy')
        if dep is None or bor is None or (dep + bor) <= 0:
            continue
        if r.get('maxRoe') is None:
            continue
        lev = round((r['maxRoe'] - dep) / (dep + bor) + 1, 2)
        rows.append(dict(r, lev=lev))
    rows.sort(key=lambda r: (0 if r['asset'] == 'ETH' else 1, -r['lev']))

    ws = wb['Looping']
    old_last = max((r for r in range(4, ws.max_row + 1)
                    if ws.cell(r, 1).value not in (None, '')), default=3)
    tmpl = {c: copy.copy(ws.cell(4, c)._style) for c in range(1, 15)}
    for r in range(4, old_last + 1):
        for c in range(1, 15):
            cell = ws.cell(r, c)
            cell.value = None
            cell.style = 'Normal'
        if r in ws.row_dimensions:
            del ws.row_dimensions[r]
    for i, d in enumerate(rows):
        r = 4 + i
        ws.row_dimensions[r].height = 15.0
        try:
            rl, rbasis = loop_risk(d, risk_scores or {})
        except Exception:
            rl, rbasis = d['risk'], ''
        vals = [d['asset'], d['protocol'], d['chain'], d['depositAsset'], d['borrowAsset'],
                d['lev'], f'=(H{r}+I{r})*(F{r}-1)+H{r}', d['depApy'], d['borApy'],
                rl, d['liquidity'], d['toTarget'], d['utilization'], d['lltv'], rbasis]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c)
            if c <= 14:
                cell._style = copy.copy(tmpl[c])
            cell.value = v
        if rl in RISK_COLOR:
            f = copy.copy(ws.cell(r, 10).font)
            f.color = openpyxl.styles.Color(rgb=RISK_COLOR[rl])
            ws.cell(r, 10).font = f
    ws.column_dimensions['O'].hidden = True
    last = 3 + len(rows)
    ws.cell(1, 1).value = (f'Looping Opportunities  ·  yieldz.io/leverage, {today}'
                           f'  ·  {len(rows)} positive-carry loops')
    ws.auto_filter.ref = f'A3:N{last}'
    ws.freeze_panes = 'C4'
    print(f'Looping: {len(rows)} rows written '
          f'({sum(1 for d in rows if d["asset"] == "ETH")} ETH, '
          f'{sum(1 for d in rows if d["asset"] == "USD")} USD)')


def restructure_farm_header(ws):
    """Rebuild the farm-tab layout: A..E as in the template, Risk in F,
    the template's F..N (Link..intrinsic) shifted to G..O, and a wide
    Risk Note column in P. The template's rating/score/pool/Holders
    columns (old O..R) are dropped. Idempotent per fresh template load."""
    # row 1 + row 2 band merges
    for rng in [str(m) for m in list(ws.merged_cells.ranges)]:
        ws.unmerge_cells(rng)
    band = {}
    for col in (9, 12):  # I2 'Total APY', L2 'Breakdown' band anchors
        c = ws.cell(2, col)
        band[col] = (c.value, copy.copy(c._style))
    for c in range(1, 20):
        ws.cell(2, c).style = 'Normal'
    for src, dst in [(9, 10), (12, 13)]:
        cell = ws.cell(2, dst)
        cell._style = band[src][1]
        cell.value = band[src][0]
    ws.merge_cells('A1:P1')
    for rng in ('J2:L2', 'M2:O2'):
        ws.merge_cells(rng)
    # row 3 headers: A..E, Risk, old Link..intrinsic, Risk Note
    hdr_style = copy.copy(ws.cell(3, 1)._style)
    hdrs = [ws.cell(3, c).value for c in range(1, 19)]
    new_hdrs = hdrs[:5] + ['Risk'] + hdrs[5:14] + ['Risk Note']
    for c in range(1, 20):
        ws.cell(3, c).style = 'Normal'
        ws.cell(3, c).value = None
    for c, v in enumerate(new_hdrs, start=1):
        cell = ws.cell(3, c)
        cell._style = copy.copy(hdr_style)
        cell.value = v
    # column widths: shift Link..intrinsic one right, Risk + Risk Note own
    old_w = {c: ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width
             for c in range(1, 19)}
    for c in range(14, 5, -1):  # old F..N -> G..O
        ws.column_dimensions[openpyxl.utils.get_column_letter(c + 1)].width = old_w.get(c)
    ws.column_dimensions['F'].width = 9.5
    ws.column_dimensions['P'].width = 70
    for col in ('Q', 'R', 'S'):
        ws.column_dimensions[col].width = 4


# ---------------------------------------------------------------------------
# Verification layer
# ---------------------------------------------------------------------------

YZ = 'https://yieldz.io/api'
CHAIN_IDS = {1: 'Ethereum', 8453: 'Base', 42161: 'Arbitrum', 10: 'OP Mainnet',
             137: 'Polygon', 56: 'BSC', 100: 'Gnosis', 59144: 'Linea',
             143: 'Monad', 999: 'Hyperliquid L1', 43114: 'Avalanche',
             5000: 'Mantle', 9745: 'Plasma', 480: 'World Chain', 146: 'Sonic'}
BLOCK_SECONDS = {1: 12, 8453: 2, 42161: 0.25, 10: 2, 137: 2.1, 56: 3,
                 100: 5, 59144: 2, 143: 0.5, 999: 1, 9745: 1, 252: 2}
CURVE_CHAINS = {'Ethereum': ('ethereum', 1), 'Arbitrum': ('arbitrum', 42161),
                'Base': ('base', 8453), 'OP Mainnet': ('optimism', 10),
                'Polygon': ('polygon', 137), 'Fraxtal': ('fraxtal', 252),
                'Gnosis': ('xdai', 100), 'BSC': ('bsc', 56),
                'Hyperliquid L1': ('hyperliquid', 999), 'Sonic': ('sonic', 146),
                'Mantle': ('mantle', 5000), 'Avalanche': ('avalanche', 43114)}
CURVE_WRAPPERS = ('Curve DEX', 'Convex Finance', 'Stake DAO', 'Beefy')
YZ_PROTO = {'Aave V3': 'aave', 'Aave V4': 'aave', 'HyperLend Pooled': 'hyperlend',
            'Fluid Lending': 'fluid', 'Euler V2': 'euler', 'Lista Lending': 'lista'}
RED_FLAGS = ('MISMATCH', 'STALE', 'NOT-ACCRUING')
FLAG_FILL = {'red': 'FFF4CCCC', 'orange': 'FFFCE5CD'}


def post_json(url, payload, timeout=30):
    req = urllib.request.Request(url, data=json.dumps(payload).encode(),
                                 headers={'Content-Type': 'application/json'})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.load(r)


def rpc(chain_id, method, params):
    return post_json(f'{YZ}/rpc/{chain_id}', dict(jsonrpc='2.0', id=1, method=method,
                                                  params=params))['result']


def history_checks(row, cut):
    """Checks that need only DefiLlama's own daily history."""
    flags = []
    raw, now = row.get('_raw') or [], row.get('now')
    ts = row.get('_last_ts')
    if ts:
        age = (datetime.datetime.now(datetime.timezone.utc)
               - datetime.datetime.fromisoformat(ts.replace('Z', '+00:00'))).days
        if age >= 2:
            flags.append(f'STALE({age}d old)')
    if len(raw) >= 10 and max(raw) == min(raw):
        flags.append('SELF-REPORTED(flat 30d history)')
    if len(raw) < 14:
        flags.append(f'NEW({len(raw)}d history)')
    if now and raw:
        m30 = statistics.mean(raw) + (row['intr'] or 0)
        if m30 > 0 and now > 3 * m30 and now - m30 > 10:
            flags.append(f'SPIKE(30d avg {m30:.1f})')
    return flags


def load_yieldz_sources():
    m = get(f'{YZ}/markets')['data']
    v = get(f'{YZ}/vaults')['data']
    lend, vaults, aave_res, mkt_by_id = {}, {}, {}, {}
    for x in m:
        proto = x['protocol'].split('_')[0]
        if proto in ('aave', 'hyperlend', 'fluid', 'euler', 'lista'):
            ch = CHAIN_IDS.get(x['chain_id'])
            key = (proto, ch, x['loan_asset']['symbol'].upper())
            lend.setdefault(key, []).append(x)
        if proto == 'aave':
            # per-reserve map so Aave collateral addresses resolve to a
            # symbol and a borrow-capacity weight (supply x LLTV)
            aave_res[(x['chain_id'], x['loan_asset']['address'].lower())] = (
                x['loan_asset']['symbol'], float(x.get('total_supply_usd') or 0),
                float(x.get('lltv') or 0))
        if proto == 'morpho' and x.get('id'):
            # market id -> collateral symbol, for look-through resolution
            cs = (x.get('collateral_asset') or {}).get('symbol')
            if cs:
                mkt_by_id[(x['chain_id'], x['id'].lower())] = cs
    for x in v:
        ch = CHAIN_IDS.get(x['chain_id'])
        vaults.setdefault((ch, x['loan_asset']['address'].lower()), []).append(x)
    return lend, vaults, aave_res, mkt_by_id


# venue quality for pooled-market allocations inside curated vaults: the
# vault supplies its own asset into these, so the venue is the exposure
VENUE_SCORE = {'aave-v3': 8.5, 'aave-v2': 7.5, 'spark': 8.2, 'spark-lend': 8.2,
               'compound-v3': 8.5, 'compound-v2': 7.0, 'moonwell': 7.5,
               'fluid': 8.0, 'fluid-instadapp': 8.0, 'gearbox': 7.0,
               'pendle': 6.5, 'harvest': 6.0, 'yearn': 7.5, 'curve': 7.5,
               'sky': 8.0, 'liquity-zapper': 7.0, 'liquity': 7.5}


# venue quality by name fragment, for scoring third-party books whose
# positions are named strings ("Morpho Vault - ...", "Sturdy crvUSD lender")
BOOK_VENUES = [('aave', 8.5), ('spark', 8.2), ('compound', 8.5), ('fluid', 8.0),
               ('curve', 7.5), ('convex', 7.0), ('balancer', 7.5), ('sky', 8.0),
               ('sturdy', 6.0), ('morpho', 6.5), ('euler', 6.5), ('ipor', 6.0),
               ('gearbox', 6.5), ('silo', 6.5), ('pendle', 6.5), ('uniswap', 7.0),
               ('aerodrome', 6.5), ('velodrome', 6.5), ('maker', 8.0),
               ('lido', 8.5), ('stakedao', 6.5), ('stake dao', 6.5),
               ('yearn', 6.5), ('lagoon', 5.9), ('termstructure', 5.5),
               ('spectra', 5.5), ('kiln', 6.0), ('gauntlet', 6.0)]


def load_morpho_vault_books():
    """Morpho curated-vault books by vault name, for resolving positions
    like 'Morpho Vault - Steakhouse Prime ETH' inside third-party books to
    the actual collateral rather than a generic venue score."""
    q = '''query V($chainId: Int!, $skip: Int!) {
      vaults(first: 100, skip: $skip,
             where: {chainId_in: [$chainId], totalAssetsUsd_gte: 1000000}) {
        items { name symbol state { allocation {
          supplyAssetsUsd market { collateralAsset { symbol } } } } } } }'''
    books = {}
    for cid in (1, 8453, 42161, 10, 137, 143, 999):
        skip = 0
        while skip < 400:
            try:
                d = get_post('https://api.morpho.org/graphql',
                             {'query': q, 'variables': {'chainId': cid, 'skip': skip}})
            except Exception:
                break
            items = ((d.get('data') or {}).get('vaults') or {}).get('items') or []
            for v in items:
                book = {}
                for a in ((v.get('state') or {}).get('allocation') or []):
                    cs = ((a.get('market') or {}).get('collateralAsset') or {}).get('symbol')
                    usd = float(a.get('supplyAssetsUsd') or 0)
                    if cs and usd > 0:
                        book[cs] = book.get(cs, 0) + usd
                if book:
                    for key in (v.get('name'), v.get('symbol')):
                        if key and len(key) >= 5:
                            books[key.lower()] = sorted(book.items(), key=lambda x: -x[1])
            if len(items) < 100:
                break
            skip += 100
    return books


_LABEL_STOP = {'usd', 'supply', 'borrow', 'vault', 'gauge', 'market', 'lender',
               'pool', 'main', 'core', 'prime', 'yield', 'basis', 'protocol',
               'eth', 'btc', 'v2', 'v3', 'lp', 'the', 'and'}


def load_addr_books():
    """Universal address -> book registry for exact position resolution:
    every Morpho curated vault (yieldz full list) resolves to its collateral
    allocation by vault address. Euler EVK vaults and Lagoon vaults are
    merged in by their loaders. Enables recursive, exact look-through for
    any strategy-vault position that carries a contract address."""
    out = {}
    d = get(f'{YZ}/vaults?include_unwhitelisted=true&protocol=morpho_vault')
    for v in d.get('data', []):
        pd_ = v.get('protocol_data') or {}
        allocs = pd_.get('current_market_allocations') or []
        book = []
        for a in allocs:
            cs = (a.get('collateral_asset') or {}).get('symbol')
            w = float(a.get('allocation_percentage') or 0)
            if cs and w > 0:
                book.append((cs, w))
        addr = (v.get('id') or '').lower()
        if addr and book:
            out[addr] = ('morpho_vault', book)
    return out


def resolve_position_addr(addr, risk_scores, depth=2):
    """Exact-identity resolution of a position by contract address against
    the universal registry. Returns (score, kind) or (None, None)."""
    if not addr or depth <= 0:
        return None, None
    ent = _MV_BOOKS.get('addr', {}).get(addr.lower())
    if not ent:
        return None, None
    kind, book = ent
    num = den = 0.0
    for item in book[:10]:
        if kind == 'lagoon':
            label, w, sub = item[0], item[1], item[2] if len(item) > 2 else None
            v = None
            if sub and sub != addr.lower():
                v, _k = resolve_position_addr(sub, risk_scores, depth - 1)
            if v is None:
                v, _d = resolve_book_position(label, risk_scores,
                                              _MV_BOOKS.get('books', {}))
            if v is None:
                v = 3.0 if (w / max(sum(i[1] for i in book), 1)) >= 0.10 else None
        else:   # collateral-symbol books (morpho_vault, euler)
            sym, w = item[0], item[1]
            v, _lbl = _collat_score(sym, risk_scores)
            if v is None:
                v = 3.0 if (w / max(sum(i[1] for i in book), 1)) >= 0.10 else None
        if v is not None:
            num += v * w
            den += w
    if den > 0:
        return num / den, kind
    return None, None


def resolve_book_position(label, risk_scores, mv_books):
    """Score one named position inside a strategy book, deepest identity
    first: a Morpho vault name resolves to that vault's actual collateral;
    else asset symbols parsed from the label score directly (worst leg);
    a recognized venue acts as a floor. Returns (score, desc) or (None,)*2."""
    low = label.lower()
    # 1. named morpho vault -> weighted collateral score of its real book
    best = None
    for name, book in mv_books.items():
        if name in low and (best is None or len(name) > len(best[0])):
            best = (name, book)
    if best:
        tot = sum(u for _, u in best[1])
        num = den = 0.0
        for cs, usd in best[1][:8]:
            v, _ = _collat_score(cs, risk_scores)
            if v is None:
                v = 3.0 if usd / tot >= 0.10 else None
            if v is not None:
                num += v * usd
                den += usd
        if den > 0:
            return num / den, f'{label[:22]}[{best[0][:14]}]'
    # 2. asset symbols in the label (pairs and single tokens)
    toks = [t for t in re.findall(r'[A-Za-z0-9+]+', label)
            if len(t) >= 3 and t.lower() not in _LABEL_STOP]
    scores = []
    for t in toks:
        v, _ = _collat_score(t, risk_scores)
        if v is not None:
            scores.append((v, t))
    venue = next((sc for k, sc in BOOK_VENUES
                  if (k in low if ' ' in k else k in set(re.findall(r'[a-z]+', low)))), None)
    if scores:
        worst = min(scores)
        v = min(worst[0], venue) if venue is not None else worst[0]
        return v, f'{label[:16]}~{worst[1]}'
    if venue is not None:
        return venue, f'{label[:22]}†'
    return None, None


def book_to_exposure(row, book, kind):
    """Convert a named-position book [(label, weight)] into an exposure
    list: known venues score at venue quality, idle/wallet entries score as
    the vault's own asset, everything else is unrated exposure."""
    out = []
    own = row['symbol'].split('-')[0]
    mv_books = _MV_BOOKS.get('books', {})
    rs = _MV_BOOKS.get('risk', {})
    for item in book:
        label, w = item[0], item[1]
        contract = item[2] if len(item) > 2 else None
        if w <= 0:
            continue
        if label.lower().startswith(('wallet', 'cash', 'idle', 'other')):
            out.append((own, w))
            continue
        # exact identity first: the position's contract address
        v, k = resolve_position_addr(contract, rs)
        if v is not None:
            out.append((f'{label[:20]}={k}', w, v))
            continue
        v, desc = resolve_book_position(label, rs, mv_books)
        if v is not None:
            out.append((desc, w, v))
        else:
            out.append((label[:26], w))      # unrated path
    if out:
        row['collat'] = out
        row['collat_kind'] = kind


# resolver context set once per run (book_to_exposure is called from deep
# inside per-row loops where threading these through is noisy)
_MV_BOOKS = {}


def load_lagoon_books():
    """Lagoon publishes each vault's live composition (protocol-level
    repartition percentages) via GraphQL — a scoreable strategy book."""
    idx = {}
    for skip in range(0, 200, 5):    # tiny pages: one vault's broken
        q = ('query { vaults(first: 5, skip: %d, where: {isVisible_eq: true}) '
             '{ items { symbol address chain { id } composition { tokenCompositions '
             '{ name repartition contract } } } } }' % skip)
        d = None
        for attempt in range(2):      # composition 500s its whole page;
            try:                      # rapid paging trips rate limits
                d = get_post('https://api.lagoon.finance/query', {'query': q})
                break
            except Exception:
                time.sleep(1.0 + attempt)
        if d is None:
            continue
        items = ((d.get('data') or {}).get('vaults') or {}).get('items') or []
        time.sleep(0.4)
        for v in items:
            comp = (v.get('composition') or {}).get('tokenCompositions') or []
            book = [(c.get('name') or '?', float(c.get('repartition') or 0),
                     (c.get('contract') or '').lower() or None)
                    for c in comp]
            try:   # lagoon serves chain ids as strings
                cid = int((v.get('chain') or {}).get('id'))
            except (TypeError, ValueError):
                continue
            ch = CHAIN_IDS.get(cid)
            if ch and book:
                idx[(ch, (v.get('symbol') or '').upper())] = book
                addr = (v.get('address') or '').lower()
                if addr:   # nested lagoon-in-lagoon resolves by address
                    _MV_BOOKS.setdefault('addr', {})[addr] = ('lagoon', book)
        if not items and skip > 0:
            break
    return idx


def get_post(url, payload):
    req = urllib.request.Request(url, data=json.dumps(payload).encode(),
                                 headers={'User-Agent': UA,
                                          'Content-Type': 'application/json'})
    try:
        with urllib.request.urlopen(req, timeout=45) as r:
            return json.load(r)
    except urllib.error.HTTPError as e:
        # graphql errors ride on 4xx/5xx but may still carry partial data
        try:
            return json.loads(e.read())
        except Exception:
            raise


def ipor_exposure(row, v, mkt_by_id, euler_by_id):
    """Look-through for a Fusion by IPOR vault: its API publishes the live
    book (marketBalances). Idle asset stays the vault asset; Morpho/Euler
    positions resolve to the market's collateral; pooled venues score by
    venue quality; anything unresolvable counts as unrated exposure."""
    hist = v.get('history') or []
    if not hist:
        return
    latest = max(hist, key=lambda h: h.get('blockNumber') or 0)
    cid = v.get('chainId')
    out = []
    for mb in latest.get('marketBalances') or []:
        bal = float(mb.get('balanceUsd') or 0)
        if bal <= 0:
            continue
        proto = (mb.get('protocol') or '').lower()
        mid = (mb.get('marketId') or '').lower()
        if proto == 'erc20':
            out.append((v.get('asset') or 'idle', bal))
        elif proto.startswith('morpho'):
            cs = mkt_by_id.get((cid, mid))
            # unresolved market: known venue, unknown collateral -> opaque 6.0
            out.append((cs, bal) if cs else (f'morpho?{mid[:8]}', bal, 6.0))
        elif proto.startswith('euler'):
            e = euler_by_id.get((cid, mid))
            if e and e.get('collat'):
                tot = sum(w for _, w in e['collat']) or 1
                out.extend((s, bal * w / tot) for s, w in e['collat'])
            else:
                v, _k = resolve_position_addr(mid, _MV_BOOKS.get('risk', {}))
                out.append((f'euler?{mid[:8]}', bal, v if v is not None else 6.0))
        elif proto in VENUE_SCORE:
            out.append((proto, bal, VENUE_SCORE[proto]))
        elif proto.endswith('erc4626'):
            # exact resolution when the wrapper address is a known vault;
            # otherwise scored like an opaque curated vault
            v, k = resolve_position_addr(mid, _MV_BOOKS.get('risk', {}))
            out.append((f'{proto}={k}' if v is not None else proto, bal,
                        v if v is not None else 6.0))
        else:
            out.append((proto or 'unknown', bal))
    if out:
        row['collat'] = out
        row['collat_kind'] = 'allocation'


def capture_exposure(row, m, aave_res):
    """Record what a matched market/vault is actually exposed to.

    - isolated pair markets (morpho/fluid/lista/euler): the collateral asset
    - aave pools: every collateral in the pool, weighted by supply x LLTV
      (its borrow capacity — a proxy for how much debt it can back)
    - curated vaults (morpho_vault): the live market allocation
    Also records utilization for the exit-liquidity deduction.
    """
    if row.get('util') is None and m.get('utilization') is not None:
        row['util'] = m['utilization']
    if row.get('collat'):
        return
    pd = m.get('protocol_data') or {}
    ca = (m.get('collateral_asset') or {}).get('symbol')
    if ca:
        row['collat'] = [(ca, 1.0)]
        row['collat_kind'] = 'collateral'
        # isolated pair market: liquidation margin + oracle quality signals
        try:
            lltv = float(m.get('lltv') or 0)
            if lltv > 100:      # basis points
                lltv /= 10000
            elif lltv > 1.5:    # percent
                lltv /= 100
            if lltv > 0:
                row['lltv'] = lltv
        except (TypeError, ValueError):
            pass
        if m.get('oracle_providers') is not None:
            row['oracle'] = m['oracle_providers']
        return
    if m.get('protocol', '').startswith('aave'):
        loan = m['loan_asset']['symbol'].upper()
        out = []
        for addr in (pd.get('borrow_yield_trends_by_collateral') or {}):
            r = aave_res.get((m['chain_id'], addr.lower()))
            if r and r[0].upper() != loan:
                w = r[1] * (r[2] / 10000)
                if w > 0:
                    out.append((r[0], w))
        if out:
            row['collat'] = out
            row['collat_kind'] = 'pool collateral'
        return
    allocs = pd.get('current_market_allocations') or []
    out = []
    for a in allocs:
        cs = (a.get('collateral_asset') or {}).get('symbol')
        w = float(a.get('allocation_percentage') or 0)
        if cs and w > 0:
            out.append((cs, w))
    if out:
        row['collat'] = out
        row['collat_kind'] = 'allocation'


def cross_source(row, lend, vaults):
    """Match a lending row to yieldz's protocol-sourced data and compare.

    Compares every APY decomposition (base, base+reward, base+intrinsic, total)
    so accounting differences between the two pipelines don't raise false
    alarms; flags only when nothing lines up.
    """
    m = None
    if row['name'] in YZ_PROTO:
        cands = lend.get((YZ_PROTO[row['name']], row['chain'], row['symbol'].upper()), [])
        if cands:
            m = max(cands, key=lambda x: x['total_supply_usd'])
    elif row['name'] == 'Morpho Blue' and row.get('_underlying'):
        cands = vaults.get((row['chain'], row['_underlying'].lower()), [])
        if cands:
            m = min(cands, key=lambda x: abs(x['total_supply_usd'] - row['tvl']))
            if not (0.5 <= (m['total_supply_usd'] + 1) / (row['tvl'] + 1) <= 2):
                m = None
    if m is None:
        return None, None, None
    src = m['supply_apy']
    # protocol-reported reward emissions (supply-side APRs, fractions)
    src_rew = sum((r.get('supply_apr') or 0) for r in (m.get('rewards') or [])) * 100
    base, rew, intr = row['base'] or 0, row['rew'] or 0, row['intr'] or 0
    # 0.6pp / 25% tolerance absorbs the intraday skew between DefiLlama's
    # snapshot and yieldz's live on-chain read
    tol = lambda a, b: abs(a - b) <= max(0.6, 0.25 * max(abs(a), abs(b)))
    comps = [('base', base, src), ('base+reward', base + rew, src),
             ('base+intrinsic', base + intr, src), ('total', base + rew + intr, src),
             ('base+reward', base + rew, src + src_rew),
             ('total', base + rew + intr, src + src_rew + intr)]
    matched = next((lbl for lbl, v, s in comps if tol(v, s)), None)
    if matched is None:
        # only accuse when the pairing is certain: with many same-asset vaults,
        # a loose TVL match may simply be the wrong vault
        ratio = (m['total_supply_usd'] + 1) / (row['tvl'] + 1)
        if not (0.85 <= ratio <= 1.18):
            return None, None, None
    return src, matched, m


# share-price read cascade: ERC4626, Yearn, Beefy, Curve, custom
SHARE_SELECTORS = ['0x07a2d13a' + hex(10 ** 18)[2:].rjust(64, '0'),  # convertToAssets(1e18)
                   '0x99530b06', '0x77c7b8fc', '0xbb7b8b80']


def onchain_realized(chain_id, addr, selector=None):
    """Annualized ~7d realized yield from on-chain share-price growth."""
    if chain_id not in BLOCK_SECONDS:
        return None
    datas = [selector] if selector else SHARE_SELECTORS
    now_block = int(rpc(chain_id, 'eth_blockNumber', []), 16)
    then_est = now_block - int(7 * 86400 / BLOCK_SECONDS[chain_id])
    hdr_now = rpc(chain_id, 'eth_getBlockByNumber', [hex(now_block), False])
    hdr_then = rpc(chain_id, 'eth_getBlockByNumber', [hex(then_est), False])
    dt = int(hdr_now['timestamp'], 16) - int(hdr_then['timestamp'], 16)
    if dt < 3 * 86400:
        return None
    for data in datas:
        try:
            pps_now = int(rpc(chain_id, 'eth_call', [{'to': addr, 'data': data}, hex(now_block)]), 16)
            pps_then = int(rpc(chain_id, 'eth_call', [{'to': addr, 'data': data}, hex(then_est)]), 16)
        except Exception:
            continue
        if pps_then > 0 and pps_now > 0:
            if pps_now == pps_then:
                # byte-identical over days is more likely a non-archive RPC
                # silently serving latest state than a frozen vault — refuse
                # to conclude either way
                return None
            return (pps_now / pps_then - 1) * (365 * 86400 / dt) * 100
    return None


def realized_flag(realized, quoted):
    if realized is None:
        return None
    if quoted > 1 and realized < 0.05:
        return f'NOT-ACCRUING(realized {realized:.2f} vs quoted {quoted:.2f})'
    if quoted > 2 and realized < 0.35 * quoted:
        return f'REALIZED-LOW({realized:.2f} vs quoted {quoted:.2f})'
    return None


def realized_check(row, m):
    """Realized-yield spot check for a yieldz-matched vault."""
    if not m or m.get('protocol') not in ('morpho_vault', 'lista_vault'):
        return None
    realized = onchain_realized(m['chain_id'], m['id'])
    if realized is None:
        return None
    return dict(realized=round(realized, 3), flag=realized_flag(realized, row['base'] or 0))


def pendle_check(row):
    """Cross-check Pendle rows against the official Pendle API.

    PT rows compare against the implied APY (flag on disagreement). LP rows
    compare against the aggregated LP APY confirm-only — semantics differ
    enough that disagreement is not evidence of bad data.
    """
    meta = (row['meta'] or '').upper().replace(' ', '')
    if row['name'] != 'Pendle':
        return None, None, None
    is_pt = 'BUYINGPT' in meta
    for mk in getattr(pendle_check, 'markets', []) or []:
        if mk['expiry'] in meta and row['symbol'].upper() in mk['sym']:
            src = mk['implied'] if is_pt else mk['aggregated']
            if src is None:
                return None, None, None
            ok = abs(src - (row['now'] or 0)) <= max(0.75, 0.25 * max(src, row['now'] or 0))
            return src, ok, ('flag' if is_pt else 'confirm-only')
    return None, None, None


def init_pendle():
    pendle_check.markets = []
    for cid in (1, 42161, 56, 143, 9745, 8453, 5000):
        try:
            skip = 0
            while True:
                res = get(f'https://api-v2.pendle.finance/core/v1/{cid}/markets'
                          f'?limit=100&skip={skip}', retries=1)
                batch = res.get('results', [])
                for mk in batch:
                    exp = (mk.get('expiry') or '')[:10]
                    if not exp:
                        continue
                    d = datetime.date.fromisoformat(exp)
                    pendle_check.markets.append(dict(
                        sym=((mk.get('pt') or {}).get('symbol') or mk.get('name') or '').upper(),
                        expiry=d.strftime('%d%b%Y').upper(),
                        implied=(mk.get('impliedApy') or 0) * 100,
                        aggregated=(mk['aggregatedApy'] * 100
                                    if mk.get('aggregatedApy') is not None else None)))
                if len(batch) < 100:
                    break
                skip += 100
        except Exception as e:
            print(f'pendle chain {cid} skipped: {type(e).__name__}')
    if pendle_check.markets:
        print(f'pendle API reachable: {len(pendle_check.markets)} markets')
    else:
        print('pendle API not reachable (allowlist api-v2.pendle.finance to enable)')


def load_curve_pools():
    """Curve pool addresses indexed by (chain, coin-symbol set) for on-chain
    virtual-price checks of Curve/Convex/Stake DAO/Beefy rows."""
    idx = {}
    for chain, (slug, chain_id) in CURVE_CHAINS.items():
        try:
            data = get(f'https://api.curve.finance/api/getPools/all/{slug}', retries=1)
            for p in data['data']['poolData']:
                coins = frozenset((c.get('symbol') or '').upper() for c in p.get('coins', []))
                if not coins:
                    continue
                crv = p.get('gaugeCrvApy') or [0, 0]
                rewards = sum((g.get('apy') or 0) for g in (p.get('gaugeRewards') or []))
                # value share of the heaviest coin, vs an equal-weight pool
                vals = []
                for c in p.get('coins', []):
                    try:
                        vals.append((c.get('symbol'),
                                     int(c['poolBalance']) / 10 ** int(c['decimals'])
                                     * (c.get('usdPrice') or 1)))
                    except (KeyError, TypeError, ValueError):
                        pass
                tot = sum(v for _, v in vals)
                imb = None
                if tot > 0 and len(vals) >= 2:
                    top_sym, top_v = max(vals, key=lambda x: x[1])
                    imb = (top_v / tot * len(vals), top_sym, top_v / tot)
                idx.setdefault((chain, coins), []).append(dict(
                    address=p['address'], usd=p.get('usdTotal') or 0, chain_id=chain_id,
                    gauge=p.get('gaugeAddress'), lp=p.get('lpTokenAddress'), imb=imb,
                    base=p.get('latestDailyApy'),
                    total_min=(p.get('latestDailyApy') or 0) + (crv[0] or 0) + rewards,
                    total_max=(p.get('latestDailyApy') or 0) + (crv[-1] or 0) + rewards))
        except Exception as e:
            print(f'curve {slug} skipped: {type(e).__name__}')
    print(f'curve API: {len(idx)} coin-set keys')
    return idx


def curve_check(row, curve_idx):
    """Curve-family verification: total-APY comparison against Curve's own
    API (base + CRV + external rewards, boost range tolerated), plus the
    on-chain virtual-price realized check when the base yield is material.
    Returns (src_value, flag)."""
    if row['name'] not in CURVE_WRAPPERS:
        return None, None, None
    coins = frozenset(s.upper() for s in row['symbol'].split('-') if s)
    cands = curve_idx.get((row['chain'], coins), [])
    if not cands:
        return None, None, None
    if row['name'] == 'Curve DEX':
        m = min(cands, key=lambda c: abs(c['usd'] - row['tvl']))
        if not (0.5 <= (m['usd'] + 1) / (row['tvl'] + 1) <= 2):
            return None, None, None
        tight = 0.8 <= (m['usd'] + 1) / (row['tvl'] + 1) <= 1.25
    else:
        tight = False  # wrapper boost differs; compare but never accuse
        if len(cands) == 1:
            m = cands[0]
        else:
            # several same-coin-set pools: identity too loose to accuse, but
            # if one pool's reward-inclusive total matches, that still
            # confirms the reward pricing
            t = row.get('now')
            if t is None:
                t = (row['base'] or 0) + (row['rew'] or 0) + (row['intr'] or 0)
            _tl = lambda a, b: abs(a - b) <= max(0.75, 0.3 * max(abs(a), abs(b)))
            m = next((c for c in cands
                      if any(_tl(t, x) for x in (c['total_min'], c['total_max']))), None)
            if m is None:
                return None, None, None
    row['_pool_addrs'] = [a for a in (m['address'], m.get('gauge'), m.get('lp')) if a]
    if m.get('imb'):
        row['pool_imb'] = m['imb']
    # reward-inclusive total comparison (Curve DEX rows only for flagging)
    total = row.get('now')
    if total is None:
        total = (row['base'] or 0) + (row['rew'] or 0) + (row['intr'] or 0)
    tol = lambda a, b: abs(a - b) <= max(0.75, 0.3 * max(abs(a), abs(b)))
    total_ok = any(tol(total, t) for t in (m['total_min'], m['total_max']))
    if total_ok and total > 0.5:
        return round(m['total_min'], 3), None, 'curve-api (total)'
    # only accuse when Curve actually reports a rate — a null/zero source is
    # missing data, not evidence
    if tight and not total_ok and total > 1 and m['total_min'] > 0.1:
        return (round(m['total_min'], 3),
                f"MISMATCH(curve total {m['total_min']:.2f})", 'curve-api (total)')
    # fall back to the realized virtual-price check on the base component
    if (row['base'] or 0) >= 1:
        realized = onchain_realized(m['chain_id'], m['address'], '0xbb7b8b80')
        if realized is not None:
            return (round(realized, 3), realized_flag(realized, row['base'] or 0),
                    'on-chain (curve vp 7d)')
    return None, None, None


# Strata tranche APRs read from their on-chain lens (found in DefiLlama's
# adapter source); getAPRs(cdo) -> (base, target, junior, senior), 1e10 scale
STRATA_LENS = '0xeA62e3a2D5FE8D5b66dc8E1bd2405AD23C851f4e'
STRATA_CDOS = {'USDE': '0x908B3921aaE4fC17191D382BB61020f2Ee6C0e20',
               'NUSD': '0x7b6c960cf185fb27ECb91c174FAe065978beDd10',
               'MHYPER': '0x39C7E67b25fB14eAec8717B20664C2E35327e6cf',
               'M1USD': '0x613D1790d9BA381D27B4071C04380Db8ED120E5f',
               'MM1USD': '0x613D1790d9BA381D27B4071C04380Db8ED120E5f',
               'USDAT': '0xa617763cEB808f43eC9D532cbE8C65819afb846b',
               'PRIME': '0xff408b4843CDD4a33CD49EB2aBe057fE8D71C234'}


def strata_check(row):
    """Compare a Strata tranche row against the protocol's on-chain lens."""
    sym = row['symbol'].upper()
    if row['name'] != 'Strata Markets' or row['chain'] != 'Ethereum':
        return None, None
    side = 'sr' if sym.startswith('SR') else ('jr' if sym.startswith('JR') else None)
    cdo = STRATA_CDOS.get(sym[2:]) if side else None
    if not cdo:
        return None, None
    data = '0x4a4b15d1' + cdo[2:].lower().rjust(64, '0')
    res = rpc(1, 'eth_call', [{'to': STRATA_LENS, 'data': data}, 'latest'])
    words = [int(res[2:][i * 64:(i + 1) * 64], 16) for i in range(4)]
    words = [w - (1 << 256) if w >= (1 << 255) else w for w in words]
    live = (words[3] if side == 'sr' else words[2]) / 1e10
    quoted = row.get('now') or 0
    ok = abs(live - quoted) <= max(0.75, 0.25 * max(live, quoted))
    return round(live, 3), ok


def load_merkl(chain_ids):
    """Live incentive-campaign APRs from Merkl (api.merkl.xyz), indexed by
    (chain_id, campaign identifier address). Returns {} while the domain is
    not allowlisted — callers treat that as 'no Merkl data'."""
    idx = {}
    try:   # single cheap probe so a blocked domain fails fast, not per-chain
        get('https://api.merkl.xyz/v4/opportunities?chainId=1&items=1', retries=1)
    except Exception:
        return {}
    for cid in chain_ids:
        page = 0
        while page < 10:
            try:
                ops = get(f'https://api.merkl.xyz/v4/opportunities'
                          f'?chainId={cid}&items=100&page={page}&status=LIVE',
                          retries=1)
            except Exception:
                break
            if not isinstance(ops, list) or not ops:
                break
            for op in ops:
                ident = (op.get('identifier') or '').lower()
                apr = op.get('apr')
                if not ident or apr is None:
                    continue
                idx.setdefault('addr', {})[(cid, ident)] = float(apr)
                proto = ((op.get('protocol') or {}).get('id') or '').lower()
                for t in op.get('tokens') or []:
                    ta = (t.get('address') or '').lower()
                    if ta:
                        idx.setdefault('contain', {}).setdefault((cid, ta), []).append(
                            dict(apr=float(apr), proto=proto))
                toks = frozenset((t.get('address') or '').lower()
                                 for t in op.get('tokens') or [])
                if len(toks) >= 2:
                    idx.setdefault('pair', {}).setdefault((cid, toks), []).append(dict(
                        apr=float(apr), tvl=float(op.get('tvl') or 0),
                        proto=((op.get('protocol') or {}).get('id') or '').lower()))
            if len(ops) < 100:
                break
            page += 1
    return idx


# row protocol -> merkl protocol-id prefix, for pair-based campaign matching
MERKL_PROTO = {'Uniswap V3': 'uniswap', 'Uniswap V4': 'uniswap',
               'Uniswap V2': 'uniswap', 'Aerodrome Slipstream': 'aerodrome',
               'Aerodrome V1': 'aerodrome', 'Velodrome V2': 'velodrome',
               'Velodrome V3': 'velodrome', 'PancakeSwap AMM': 'pancake',
               'Curve DEX': 'curve', 'Balancer V2': 'balancer',
               'Balancer V3': 'balancer', 'Camelot V2': 'camelot',
               'Camelot V3': 'camelot', 'Fluid DEX': 'fluid',
               'Joe V2.1': 'traderjoe', 'Joe V2.2': 'traderjoe',
               'Aave V3': 'aave', 'Aave V4': 'aave', 'Morpho Blue': 'morpho',
               'Euler V2': 'euler', 'Curvance': 'curvance', 'Mento V3': 'mento',
               'Neverland': 'neverland', 'Sky Lending': 'sky',
               'Dolomite': 'dolomite', 'Curve LlamaLend': 'curve',
               'Lista Lending': 'lista', 'Fusion by IPOR': 'ipor',
               'Compound V3': 'compound', 'SparkLend': 'spark',
               'Upshift': 'upshift', 'Fluid Lending': 'fluid'}


def merkl_reward_match(row, merkl_idx):
    """Confirm a row's emission APR against a Merkl campaign.

    Identity, strongest first: exact address (registry vault, matched Curve
    pool/gauge/LP token, underlying), then same-protocol-family campaigns on
    the exact token pair (TVL-gated when several fee tiers exist). Returns
    the campaign APR when it corroborates DL's reward figure."""
    cid = next((k for k, v in CHAIN_IDS.items() if v == row['chain']), None)
    if cid is None:
        return None
    rew = row.get('rew') or 0
    close = lambda a: abs(a - rew) <= max(1.0, 0.35 * max(a, rew))
    addrs = [row.get('_registry_addr'), row.get('_underlying')]
    addrs += row.get('_pool_addrs') or []
    for addr in addrs:
        if addr:
            apr = merkl_idx.get('addr', {}).get((cid, addr.lower()))
            if apr is not None and close(apr):
                return apr
    fam = MERKL_PROTO.get(row['name'])
    # lending rows: any same-family campaign involving the underlying whose
    # APR corroborates DL's reward figure
    if fam and row.get('_underlying'):
        for c in merkl_idx.get('contain', {}).get((cid, row['_underlying'].lower()), []):
            if c['proto'].startswith(fam) and close(c['apr']):
                return c['apr']
    unders = [a for a in (row.get('_underlyings') or []) if a]
    if fam and len(unders) >= 2:
        key = (cid, frozenset(a.lower() for a in unders))
        cands = [c for c in merkl_idx.get('pair', {}).get(key, [])
                 if c['proto'].startswith(fam)]
        if len(cands) > 1:   # several fee tiers: require the same pool size
            cands = [c for c in cands if c['tvl']
                     and 0.3 <= (c['tvl'] + 1) / ((row.get('tvl') or 0) + 1) <= 3]
        for c in cands:
            if close(c['apr']):
                return c['apr']
    return None


def run_verification(tabs, registry=None):
    try:
        lend, vaults, aave_res, mkt_by_id = load_yieldz_sources()
    except Exception as e:
        print('yieldz sources unavailable, cross-source skipped:', e)
        lend, vaults, aave_res, mkt_by_id = {}, {}, {}, {}
    init_pendle()
    registry = registry or {}
    try:
        curve_idx = load_curve_pools()
    except Exception as e:
        print('curve API unavailable:', e)
        curve_idx = {}
    proto_src = {}
    try:
        import os
        import sys as _sys
        _sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        import protocol_checks
        proto_src = protocol_checks.load_protocol_sources()
    except Exception as e:
        print('protocol checks unavailable:', type(e).__name__, e)
        protocol_checks = None
    lagoon_books = {}
    try:
        lagoon_books = load_lagoon_books()
        print(f'lagoon: {len(lagoon_books)} vault books')
    except Exception as e:
        print('lagoon books unavailable:', type(e).__name__)
    try:
        _MV_BOOKS['books'] = load_morpho_vault_books()
        print(f"morpho vault books: {len(_MV_BOOKS['books'])} names")
    except Exception as e:
        print('morpho vault books unavailable:', type(e).__name__)
    merkl_idx = {}
    try:
        merkl_idx = load_merkl([c for c in CHAIN_IDS if c != 0])
        if merkl_idx:
            print(f"merkl: {len(merkl_idx.get('addr', {}))} live campaigns, "
                  f"{len(merkl_idx.get('pair', {}))} token-pair keys")
        else:
            print('merkl unavailable (api.merkl.xyz not allowlisted?) — '
                  'reward APRs unverifiable beyond yieldz/curve')
    except Exception as e:
        print('merkl unavailable:', type(e).__name__)
    try:
        _MV_BOOKS.setdefault('addr', {}).update(load_addr_books())
        print(f"addr books: {len(_MV_BOOKS['addr'])} contracts")
    except Exception as e:
        print('addr books unavailable:', type(e).__name__)
    # euler vault id -> entry (with collateral allocations), for look-through
    euler_by_id = {}
    for cands in (proto_src.get('euler') or {}).values():
        for e in cands:
            if e.get('address'):
                euler_by_id[(e.get('chain_id'), e['address'].lower())] = e
                if e.get('collat'):
                    _MV_BOOKS.setdefault('addr', {})[e['address'].lower()] = (
                        'euler', e['collat'])
    counts = collections.Counter()
    for tab, rows in tabs.items():
        cut = CUTOFF[tab.split()[0]]
        for row in rows:
            flags = history_checks(row, cut)
            src_val = src_name = None
            try:
                src, matched, m = cross_source(row, lend, vaults)
                if m is not None:
                    capture_exposure(row, m, aave_res)
                if m is not None and str(m.get('id', '')).startswith('0x'):
                    row.setdefault('_pool_addrs', []).append(m['id'])
                if src is not None:
                    src_val = round(src, 3)
                    src_name = f'yieldz ({matched})' if matched else 'yieldz'
                    if not matched:
                        flags.append(f'MISMATCH(source {src:.2f})')
                    else:
                        # a reward-inclusive decomposition matching means the
                        # protocol's own reward pricing agrees with DL's
                        if matched in ('base+reward', 'total') and (row['rew'] or 0) > 0:
                            row['_rew_confirmed'] = 'yieldz'
                        r = realized_check(row, m)
                        if r and r['flag']:
                            flags.append(r['flag'])
            except Exception:
                pass
            if src_val is None:
                try:
                    src, ok, mode = pendle_check(row)
                    if src is not None and (ok or mode == 'flag'):
                        src_val, src_name = round(src, 3), 'pendle-api'
                        if not ok:
                            flags.append(f'MISMATCH(source {src:.2f})')
                except Exception:
                    pass
            # a live source confirming the value supersedes the staleness flag
            if src_val is not None and not any(f.startswith('MISMATCH') for f in flags):
                flags = [f for f in flags if not f.startswith('STALE')]
            if src_val is None and curve_idx:
                try:
                    cv, fl, lbl = curve_check(row, curve_idx)
                    if cv is not None:
                        src_val, src_name = cv, lbl
                        if fl:
                            flags.append(fl)
                        elif (row['rew'] or 0) > 0:
                            # curve totals include gauge CRV + extra rewards
                            row['_rew_confirmed'] = 'curve gauge'
                except Exception:
                    pass
            if src_val is None and row['name'] == 'Strata Markets':
                try:
                    sv, ok = strata_check(row)
                    if sv is not None:
                        src_val, src_name = sv, 'strata on-chain lens'
                        if not ok:
                            flags.append(f'MISMATCH(source {sv:.2f})')
                except Exception:
                    pass
            if src_val is None and proto_src:
                try:
                    pv, ok, lbl, addr = protocol_checks.check_row(row, proto_src)
                    if addr and row['pool_id'] not in registry:
                        # runtime registry entry so the realized check can run
                        cid = next((k for k, v in CHAIN_IDS.items() if v == row['chain']), None)
                        if cid:
                            registry[row['pool_id']] = dict(chain_id=cid, address=addr)
                    if pv is not None:
                        src_val, src_name = pv, lbl
                        if not ok:
                            flags.append(f'MISMATCH(source {pv:.2f})')
                except Exception:
                    pass
            if src_val is None and row['pool_id'] in registry:
                try:
                    ent = registry[row['pool_id']]
                    realized = onchain_realized(ent['chain_id'], ent['address'],
                                                ent.get('selector'))
                    if realized is not None:
                        src_val = round(realized, 3)
                        src_name = 'on-chain (7d realized)'
                        fl = realized_flag(realized, row['base'] or 0)
                        if fl:
                            flags.append(fl)
                except Exception:
                    pass
            # curvance pair markets: the manager's other token is the collateral
            if (not row.get('collat') and row['name'] == 'Curvance'
                    and '/' in (row['meta'] or '')):
                others = [s for s in row['meta'].split('/')
                          if risk_norm(s) != risk_norm(row['symbol'])]
                if others:
                    row['collat'] = [(s, 1.0) for s in others]
                    row['collat_kind'] = 'collateral'
            # ipor fusion vaults publish their live book — look through it
            if (not row.get('collat') and row['name'] == 'Fusion by IPOR'
                    and proto_src.get('ipor')):
                try:
                    v = proto_src['ipor'].get((row.get('meta') or '').strip().lower())
                    if v:
                        ipor_exposure(row, v, mkt_by_id, euler_by_id)
                except Exception:
                    pass
            # yearn: ydaemon strategy list stashed during the cross-check
            if not row.get('collat') and row.get('_strategy_book'):
                book_to_exposure(row, row['_strategy_book'], 'strategy book')
            # lagoon: composition published via graphql
            if (not row.get('collat') and row['name'] == 'Lagoon' and lagoon_books):
                book = lagoon_books.get((row['chain'], row['symbol'].upper()))
                if book:
                    book_to_exposure(row, book, 'strategy book')
            # emissions-dominated APY whose reward pricing no independent
            # source confirmed (yieldz reward decomposition, curve gauge
            # totals, or Merkl campaign APRs all count as confirmation)
            total = (row['base'] or 0) + (row['rew'] or 0) + (row['intr'] or 0)
            if (row['rew'] or 0) > 0.66 * total and total > 1:
                if not row.get('_rew_confirmed') and merkl_idx:
                    row['_registry_addr'] = (registry.get(row['pool_id']) or {}).get('address')
                    mk = merkl_reward_match(row, merkl_idx)
                    if mk is not None:
                        row['_rew_confirmed'] = 'merkl'
                        if src_val is None:
                            src_val, src_name = round(mk, 3), 'merkl (reward APR)'
                if not row.get('_rew_confirmed'):
                    flags.append(f'REWARD-HEAVY(emissions {row["rew"]:.1f} of {total:.1f} unverified)')
            row['flags'] = flags
            row['src_val'], row['src_name'] = src_val, src_name
            if any(f.startswith(RED_FLAGS) for f in flags):
                row['verdict'] = 'FLAGGED'
            elif flags:
                row['verdict'] = 'CAUTION'
            elif src_val is not None:
                row['verdict'] = 'VERIFIED'
            else:
                row['verdict'] = 'UNVERIFIED'
            counts[row['verdict']] += 1
    print('verification:', dict(counts))


def build_check_tab(wb, tabs, today):
    if 'Data Check' in wb.sheetnames:
        del wb['Data Check']
    ws = wb.create_sheet('Data Check')
    header = ['Tab', 'Farm', 'Chain', 'Protocol', 'APY Now', '7d', 'Verdict',
              'Cross-Source APY', 'Source', 'Flags', 'Risk', 'Risk Basis']
    ws.cell(1, 1).value = (f'Data Check  ·  {today}  ·  history checks: all rows; '
                           f'cross-source: yieldz protocol data + on-chain reads')
    ws.cell(1, 1).font = openpyxl.styles.Font(bold=True, size=12, color='FF1F3552', name='Arial')
    for c, h in enumerate(header, start=1):
        cell = ws.cell(2, c)
        cell.value = h
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFFFF', size=10, name='Arial')
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='FF1F3552')
    order = {'FLAGGED': 0, 'CAUTION': 1, 'UNVERIFIED': 2, 'VERIFIED': 3}
    allrows = [(tab, r) for tab, rows in tabs.items() for r in rows]
    allrows.sort(key=lambda x: (order.get(x[1].get('verdict'), 9), -(x[1].get('now') or 0)))
    for i, (tab, r) in enumerate(allrows):
        vals = [tab, (r.get('farm_disp') or r.get('meta') or r['symbol']), r['chain'], r['name'],
                r.get('now'), r.get('d7'), r['verdict'], r.get('src_val'),
                r.get('src_name'), '; '.join(r.get('flags') or []),
                (f"{r['risk']} ({r['risk_score']})" if r.get('risk_score') is not None
                 else r.get('risk')),
                r.get('risk_basis')]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(3 + i, c)
            cell.value = v
            cell.font = openpyxl.styles.Font(size=10, name='Arial')
        col = ('FFC00000' if r['verdict'] == 'FLAGGED' else
               'FFB45F06' if r['verdict'] == 'CAUTION' else
               'FF1E7145' if r['verdict'] == 'VERIFIED' else 'FF7F7F7F')
        ws.cell(3 + i, 7).font = openpyxl.styles.Font(size=10, name='Arial', bold=True, color=col)
        if r.get('risk') in RISK_COLOR:
            ws.cell(3 + i, 11).font = openpyxl.styles.Font(
                size=10, name='Arial', bold=True, color=RISK_COLOR[r['risk']])
    for col, w in zip('ABCDEFGHIJKL', [10, 38, 12, 18, 9, 9, 12, 15, 14, 45, 9, 45]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A3'
    print(f'Data Check: {len(allrows)} rows')


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2],
         sys.argv[3] if len(sys.argv) > 3 else None,
         sys.argv[4] if len(sys.argv) > 4 else None)
